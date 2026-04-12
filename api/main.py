"""XLSForm Debugger v2 — FastAPI backend for converting XLSForm to ODK XForm XML."""

import base64
import csv
import io
import json
import os
import shutil
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Optional
from xml.sax.saxutils import escape as xml_escape

import openpyxl
import openpyxl.utils.exceptions
import pyxform
import pyxform.xls2xform
import pyxform.errors
import uvicorn
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

app = FastAPI(title="XLSForm Debugger v2 API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:5174"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def _save_upload(upload: UploadFile, dest: Path) -> Path:
    filepath = dest / upload.filename
    with open(filepath, "wb") as f:
        f.write(upload.file.read())
    return filepath


def _convert_xlsform(
    xlsform_path: Path, tmp_dir: Path
) -> tuple[str, list[str], str, str]:
    xform_path = tmp_dir / "output.xml"

    warnings = pyxform.xls2xform.xls2xform_convert(
        xlsform_path=str(xlsform_path),
        xform_path=str(xform_path),
        validate=False,
    )
    warnings = warnings or []

    xform_xml = xform_path.read_text(encoding="utf-8")

    title = ""
    form_id = ""
    try:
        root = ET.fromstring(xform_xml)
        ns = {"h": "http://www.w3.org/1999/xhtml", "x": "http://www.w3.org/2002/xforms"}
        title_el = root.find(".//h:head/h:title", ns)
        if title_el is not None and title_el.text:
            title = title_el.text
        instance = root.find(".//{http://www.w3.org/2002/xforms}instance")
        if instance is not None and len(instance) > 0:
            form_id = instance[0].attrib.get("id", "")
    except ET.ParseError:
        pass

    return xform_xml, list(warnings), title, form_id


def _csv_to_xml(csv_content: str, filename: str) -> dict[str, str]:
    """Convert CSV content to enketo external data XML format.
    
    Sanitizes column names to valid XML tags.
    Also injects a plain <label> element from the first English label column
    so enketo itemsets with ref="label" work correctly.
    """
    import re as _re
    bom_stripped = csv_content.lstrip("\ufeff").lstrip("\ufffe")
    reader = csv.DictReader(io.StringIO(bom_stripped))
    fieldnames = reader.fieldnames or []
    
    # Find the label source column
    label_col = None
    if "label" in fieldnames:
        label_col = "label"
    else:
        for f in fieldnames:
            if f.lower().startswith("label::english") or f.lower() == "label":
                label_col = f
                break
        if not label_col:
            for f in fieldnames:
                if "label" in f.lower():
                    label_col = f
                    break

    items = []
    for row in reader:
        item_parts = []
        has_label = False
        for k, v in row.items():
            raw_key = k.strip().lstrip("\ufeff")
            safe_key = _re.sub(r"[^a-zA-Z0-9_]", "_", raw_key)
            if not safe_key or safe_key[0].isdigit():
                safe_key = "col_" + safe_key
            item_parts.append(f"<{safe_key}>{xml_escape(str(v))}</{safe_key}>")
            if safe_key == "label":
                has_label = True
        # Inject plain <label> from English label column if not already present
        if not has_label and label_col and label_col in row:
            item_parts.append(f"<label>{xml_escape(str(row[label_col]))}</label>")
        items.append("<item>" + "".join(item_parts) + "</item>")
    file_id = Path(filename).stem
    return {"id": file_id, "xml": "<root>" + "".join(items) + "</root>"}


def _read_xlsx_rows(xlsx_path: Path) -> dict[str, list[dict[str, str]]]:
    """Read survey, choices, settings sheets from an xlsx file as lists of row dicts."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    result: dict[str, list[dict[str, str]]] = {}
    for sheet_name in ("survey", "choices", "settings"):
        if sheet_name not in wb.sheetnames:
            result[sheet_name] = []
            continue
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h or "").strip() for h in next(rows_iter)]
        except StopIteration:
            result[sheet_name] = []
            continue
        sheet_rows: list[dict[str, str]] = []
        for row in rows_iter:
            row_dict = {
                headers[i]: str(cell) if cell is not None else ""
                for i, cell in enumerate(row)
                if i < len(headers) and headers[i]
            }
            if any(v for v in row_dict.values()):
                sheet_rows.append(row_dict)
        result[sheet_name] = sheet_rows
    wb.close()
    return result


def _write_xlsx_from_rows(
    survey: list[dict[str, str]],
    choices: list[dict[str, str]],
    settings: list[dict[str, str]],
    dest: Path,
) -> Path:
    """Build an xlsx file from row dicts for survey/choices/settings sheets."""
    wb = openpyxl.Workbook()
    for idx, (sheet_name, rows) in enumerate(
        [("survey", survey), ("choices", choices), ("settings", settings)]
    ):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = sheet_name
        if not rows:
            continue
        headers = list(rows[0].keys())
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        for row_idx, row in enumerate(rows, 2):
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
    filepath = dest / "rebuilt.xlsx"
    wb.save(str(filepath))
    wb.close()
    return filepath


class ConvertJsonRequest(BaseModel):
    survey: list[dict[str, str]]
    choices: list[dict[str, str]]
    settings: list[dict[str, str]]


_MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10 MB


@app.post("/convert")
async def convert(
    xlsx_file: UploadFile = File(...),
    csv_files: Optional[list[UploadFile]] = File(None),
):
    # Validate extension
    filename = xlsx_file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    # Validate file size
    contents = await xlsx_file.read()
    if len(contents) > _MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=400, detail="File too large (max 10 MB)")
    await xlsx_file.seek(0)

    tmp_dir = Path(tempfile.mkdtemp())
    try:
        xlsform_path = _save_upload(xlsx_file, tmp_dir)

        external_data: list[dict[str, str]] = []
        if csv_files:
            for csv_file in csv_files:
                _save_upload(csv_file, tmp_dir)
                csv_file.file.seek(0)
                csv_content = csv_file.file.read().decode("utf-8")
                external_data.append(_csv_to_xml(csv_content, csv_file.filename or "data.csv"))

        xform_xml, warnings, title, form_id = _convert_xlsform(xlsform_path, tmp_dir)
        xls_rows = _read_xlsx_rows(xlsform_path)

        return {
            "xform_xml": xform_xml,
            "warnings": warnings,
            "title": title,
            "id": form_id,
            "external_data": external_data,
            "survey": xls_rows.get("survey", []),
            "choices": xls_rows.get("choices", []),
            "settings": xls_rows.get("settings", []),
        }
    except pyxform.errors.PyXFormError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except openpyxl.utils.exceptions.InvalidFileException as exc:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {exc}")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Internal error: {exc}")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@app.post("/convert-json")
async def convert_json(body: ConvertJsonRequest):
    """Convert edited xlsRows (JSON) back to XForm XML."""
    tmp_dir = Path(tempfile.mkdtemp())
    try:
        xlsx_path = _write_xlsx_from_rows(
            body.survey, body.choices, body.settings, tmp_dir
        )
        xform_xml, warnings, title, form_id = _convert_xlsform(xlsx_path, tmp_dir)
        return {
            "xform_xml": xform_xml,
            "warnings": warnings,
            "title": title,
            "id": form_id,
            "survey": body.survey,
            "choices": body.choices,
            "settings": body.settings,
        }
    except pyxform.errors.PyXFormError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Internal error: {exc}")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@app.post("/validate")
async def validate(
    xlsx_file: UploadFile = File(...),
    csv_files: Optional[list[UploadFile]] = File(None),
):
    tmp_dir = Path(tempfile.mkdtemp())
    try:
        xlsform_path = _save_upload(xlsx_file, tmp_dir)

        if csv_files:
            for csv_file in csv_files:
                _save_upload(csv_file, tmp_dir)

        _xform_xml, warnings, _title, _form_id = _convert_xlsform(xlsform_path, tmp_dir)

        return {"valid": True, "errors": [], "warnings": warnings}
    except pyxform.errors.PyXFormError as exc:
        return {"valid": False, "errors": [str(exc)], "warnings": []}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Internal error: {exc}")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@app.get("/health")
async def health():
    return {"status": "ok", "pyxform_version": pyxform.__version__}


from fastapi.responses import Response


class ExportRequest(BaseModel):
    survey: list[dict[str, str]]
    choices: list[dict[str, str]]
    settings: list[dict[str, str]]
    filename: str = "form.xlsx"


@app.post("/export")
async def export_xlsx(body: ExportRequest):
    """Export survey/choices/settings rows as a downloadable .xlsx file."""
    tmp_dir = Path(tempfile.mkdtemp())
    try:
        xlsx_path = _write_xlsx_from_rows(
            body.survey, body.choices, body.settings, tmp_dir
        )
        xlsx_bytes = xlsx_path.read_bytes()
        safe_filename = body.filename if body.filename.endswith(".xlsx") else body.filename + ".xlsx"
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Export failed: {exc}")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


class DeployRequest(BaseModel):
    survey: list[dict[str, str]]
    choices: list[dict[str, str]]
    settings: list[dict[str, str]]
    kobo_url: str = "https://kf.kobotoolbox.org"
    api_token: str = ""
    form_name: str = "New Form"


@app.post("/deploy")
async def deploy_to_kobo(body: DeployRequest):
    """Deploy form to KoboToolbox via REST API.

    Builds .xlsx, then POSTs to KoboToolbox /api/v2/assets/ endpoint.
    """
    if not body.api_token:
        raise HTTPException(status_code=400, detail="API token is required")

    tmp_dir = Path(tempfile.mkdtemp())
    try:
        xlsx_path = _write_xlsx_from_rows(
            body.survey, body.choices, body.settings, tmp_dir
        )
        xlsx_bytes = xlsx_path.read_bytes()

        import requests
        kobo_base = body.kobo_url.rstrip("/")
        resp = requests.post(
            f"{kobo_base}/api/v2/assets/",
            headers={"Authorization": f"Token {body.api_token}"},
            data={"name": body.form_name, "asset_type": "survey"},
            files={"file": (f"{body.form_name}.xlsx", xlsx_bytes,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30,
        )
        if resp.status_code >= 400:
            raise HTTPException(
                status_code=resp.status_code,
                detail=f"KoboToolbox API error: {resp.text[:500]}",
            )
        return resp.json()
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Deploy failed: {exc}")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=5050)
